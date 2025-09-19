import os, sqlite3, hashlib, json
from datetime import date, datetime, timedelta
from calendar import monthrange
from typing import Optional, List, Dict, Any, Union
from pathlib import Path

# Constants
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "leave_data.sqlite")
HOLIDAYS_JSON = os.path.join(BASE_DIR, "holidays.json")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Core Functions
def init_db():
    """Initialize database with all required tables"""
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("PRAGMA foreign_keys = ON;")
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS Users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            login_name TEXT UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            address TEXT,
            phone TEXT,
            rest_days_odd TEXT,
            rest_days_even TEXT,
            TotalLeave REAL DEFAULT 0,
            SickLeave REAL DEFAULT 0,
            cultivationLeave REAL DEFAULT 0,
            compassionateLeave REAL DEFAULT 0,
            hospital_leave REAL DEFAULT 0,
            ReplacementLeave REAL DEFAULT 0,
            cf_leave REAL DEFAULT 0,
            years_worked INTEGER DEFAULT 0,
            upd TEXT DEFAULT NULL,
            expired REAL DEFAULT 0,
            pregnantLeave REAL DEFAULT 0
        );""")
        c.execute("""CREATE TABLE IF NOT EXISTS LeaveRequests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            leave_type TEXT NOT NULL,
            start_date TEXT NOT NULL,
            num_days REAL NOT NULL,
            notes TEXT,
            status TEXT DEFAULT 'Pending',
            replacement_leave TEXT,
            leave_source TEXT,
            attachment_path TEXT,
            attachment_blob BLOB,
            bal REAL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES Users(id) ON DELETE CASCADE
        );""")
        c.execute("""CREATE TABLE IF NOT EXISTS LeaveUpdateLog (
            last_updated TEXT,
            march_processed INTEGER DEFAULT 0
        );""")
        c.execute("""CREATE TABLE IF NOT EXISTS LeaveSnapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            TotalLeave REAL, SickLeave REAL, cultivationLeave REAL,
            compassionateLeave REAL, hospital_leave REAL, ReplacementLeave REAL,
            cf_leave REAL, pregnantLeave REAL
        );""")
        c.execute("""CREATE TABLE IF NOT EXISTS LeaveAudit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT,
            performed_by TEXT,
            target_user INTEGER,
            target_request_id INTEGER,
            change_summary TEXT,
            full_diff TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        conn.commit()
    ensure_default_admin()

def ensure_default_admin():
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT id FROM Users WHERE login_name = 'admin' OR username = 'admin'")
        if not c.fetchone():
            pwd = hashlib.sha256("admin".encode()).hexdigest()
            c.execute("INSERT INTO Users (username, login_name, password_hash, role) VALUES (?,?,?,?)",
                      ("admin","admin", pwd, "superadmin"))
            conn.commit()

def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def get_holidays() -> Dict[str, Dict[str, str]]:
    """Get all holidays from the JSON file"""
    try:
        with open(HOLIDAYS_JSON, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        # Return empty structure if file doesn't exist or is invalid
        return {"defaults": {}, str(datetime.now().year): {}}

def verify_login(login_name: str, password: str):
    pwd = hash_password(password)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM Users WHERE login_name = ? AND password_hash = ?", (login_name, pwd))
        r = c.fetchone()
        if not r:
            return None
        return dict(r)

def create_user(username: str, login_name: str, password: str, role: str="user") -> int:
    pwd = hash_password(password)
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("INSERT INTO Users (username, login_name, password_hash, role) VALUES (?,?,?,?)",
                  (username, login_name, pwd, role))
        conn.commit()
        return c.lastrowid


def get_user_leaves(user_id: int):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM LeaveRequests WHERE user_id = ? ORDER BY timestamp DESC", (user_id,))
        return [dict(r) for r in c.fetchall()]


def submit_leave(user_id: int, leave_type: str, start_date: str, num_days: float, notes: str=None, replacement_leave: str=None, attachment_bytes: bytes=None, attachment_filename: str=None) -> int:
    """Submit a leave request. For multiple dates, start_date should contain dates separated by semicolons."""
    # Handle multiple dates
    dates = [d.strip() for d in start_date.split(';') if d.strip()]
    if len(dates) > 1:
        # For multiple dates, create separate entries
        leave_ids = []
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            for date in dates:
                c.execute(""" 
                    INSERT INTO LeaveRequests 
                    (user_id, leave_type, start_date, num_days, notes, replacement_leave, attachment_image) 
                    VALUES (?,?,?,?,?,?,?)
                """, (user_id, leave_type, date, 1, notes, replacement_leave, attachment_bytes))
                leave_id = c.lastrowid
                leave_ids.append(leave_id)
                audit("Submit Leave", 
                      performed_by=str(user_id), 
                      target_user=user_id, 
                      target_request_id=leave_id, 
                      change_summary=f"{leave_type} {date} 1d")
            conn.commit()
        return leave_ids[0]  # Return first leave ID
    else:
        # Single date handling (original logic)
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            c.execute(""" 
                INSERT INTO LeaveRequests 
                (user_id, leave_type, start_date, num_days, notes, replacement_leave, attachment_image) 
                VALUES (?,?,?,?,?,?,?)
            """, (user_id, leave_type, start_date, num_days, notes, replacement_leave, attachment_bytes))
            conn.commit()
            lrid = c.lastrowid
            audit("Submit Leave", 
                  performed_by=str(user_id), 
                  target_user=user_id, 
                  target_request_id=lrid, 
                  change_summary=f"{leave_type} {start_date} {num_days}d")
            return lrid

def get_pending_leaves():
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT lr.id, u.username, lr.leave_type, lr.start_date, lr.num_days, lr.notes FROM LeaveRequests lr JOIN Users u ON lr.user_id = u.id WHERE lr.status = 'Pending' ORDER BY lr.timestamp ASC")
        return [dict(r) for r in c.fetchall()]


def audit(action: str, performed_by: str, target_user: int=None, target_request_id: int=None, change_summary: str="", full_diff: dict=None):
    if isinstance(full_diff, dict):
        full_diff = json.dumps(full_diff, ensure_ascii=False)
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("INSERT INTO LeaveAudit (action, performed_by, target_user, target_request_id, change_summary, full_diff) VALUES (?,?,?,?,?,?)",
                  (action, performed_by, target_user, target_request_id, change_summary, full_diff))
        conn.commit()

def get_audit_logs(limit: int=200):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT id, action, performed_by, target_user, target_request_id, change_summary, full_diff, timestamp FROM LeaveAudit ORDER BY timestamp DESC LIMIT ?", (limit,))
        return c.fetchall()


def apply_annual_leave_updates():
    today = date.today()

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()

        # === Init LeaveUpdateLog if needed ===
        c.execute("SELECT last_updated, march_processed FROM LeaveUpdateLog")
        row = c.fetchone()
        if row:
            last_updated = date.fromisoformat(row[0])
            march_transfer_done = bool(row[1])
        else:
            last_updated = date(2000, 1, 1)
            march_transfer_done = False
            c.execute("INSERT INTO LeaveUpdateLog (last_updated, march_processed) VALUES (?, 0)",
                      (last_updated.isoformat(),))

        # === Skip if already updated this month ===
        if last_updated.year == today.year and last_updated.month == today.month:
            print("Leave already updated this month.")
            return

        # === Check yearly reset & March expiry ===
        is_new_year = today.year != last_updated.year
        crossed_march = today.month > 3 and not march_transfer_done

        holidays_data = load_holiday_json()
        current_year = str(today.year)
        month_holidays = {
            datetime.strptime(date_str, "%Y-%m-%d").date()
            for date_str in holidays_data.get(current_year, {})
            if datetime.strptime(date_str, "%Y-%m-%d").month == today.month
        }

        start_month = date(today.year, today.month, 1)
        days_in_month = monthrange(today.year, today.month)[1]

        # === Loop users ===
        c.execute("SELECT id, username, years_worked, TotalLeave, cf_leave FROM Users")
        users = c.fetchall()

        for user_id, username, years_worked, current_total, cf_old in users:
            expired = {}

            # === Reset in new year ===
            if is_new_year:
                c.execute("""
                    SELECT TotalLeave, SickLeave, cultivationLeave,
                           compassionateLeave, hospital_leave,
                           ReplacementLeave, pregnantLeave
                    FROM Users WHERE id = ?
                """, (user_id,))
                t, s, x, comp, hosp, rep, preg = c.fetchone()

                expired = {
                    "TotalLeave": t,
                    "SickLeave": s,
                    "cultivationLeave": x,
                    "compassionateLeave": comp,
                    "hospital_leave": hosp,
                    "ReplacementLeave": rep,
                    "pregnantLeave": preg
                }

                # Reset & top-up logic
                sick_leave = 18 if years_worked > 5 else 12
                cf_new = min(current_total, 5)
                c.execute("""
                    UPDATE Users SET
                        cf_leave = ?, TotalLeave = 0, SickLeave = ?,
                        hospital_leave = 60, cultivationLeave = 7,
                        compassionateLeave = 14, ReplacementLeave = 0, pregnantLeave = 98,
                        years_worked = years_worked + 1
                    WHERE id = ?
                """, (cf_new, sick_leave, user_id))
                current_total = 0
                conn.commit()
                # Audit log
                log_action(
                    action="Yearly Reset",
                    performed_by="System",
                    target_user=user_id,
                    change_summary=f"Expired: {expired}, CF set to {cf_new}, SickLeave set to {sick_leave}",
                    full_diff=expired,
                    conn=conn
                )

            # === Snapshot before March CF clears ===
            if crossed_march:
                expired["cf_leave"] = cf_old
                c.execute("UPDATE Users SET cf_leave = 0 WHERE id = ?", (user_id,))
                conn.commit()
                # Audit log
                log_action(
                    action="March Expiry",
                    performed_by="System",
                    target_user=user_id,
                    change_summary=f"CF leave {cf_old} expired and set to 0",
                    full_diff={"cf_leave": cf_old},
                    conn=conn
                )

            # === Insert into LeaveSnapshots if any expired ===
            if expired:
                cols = ', '.join(['user_id'] + list(expired.keys()))
                placeholders = ', '.join(['?'] * (1 + len(expired)))
                values = [user_id] + list(expired.values())
                c.execute(f"""
                    INSERT INTO LeaveSnapshots ({cols})
                    VALUES ({placeholders})
                """, values)

            # === Monthly Top-up ===
            if years_worked > 10:
                add_total = 1.58334
            elif years_worked > 5:
                add_total = 1.5
            else:
                add_total = 1.0

            new_total = round(current_total + add_total, 2)
            c.execute("UPDATE Users SET TotalLeave = ? WHERE id = ?", (new_total, user_id))
            conn.commit()
            log_action(
                action="Monthly Top-up",
                performed_by="System",
                target_user=user_id,
                change_summary=f"Added {add_total} days. New TotalLeave: {new_total}",
                full_diff={"add_days": add_total, "new_TotalLeave": new_total},
                conn=conn
            )

            # === Bonus for OFF-day holiday ===
            for offset in range(days_in_month):
                day = start_month + timedelta(days=offset)
                if day in month_holidays:
                    rest_days = [r.lower() for r in get_alternating_rest_days(username, day)]
                    if day.strftime("%A").lower() in rest_days:
                        c.execute("UPDATE Users SET TotalLeave = TotalLeave + 1 WHERE id = ?", (user_id,))
                        conn.commit()
                        log_action(
                            action="Bonus Off-Day",
                            performed_by="System",
                            target_user=user_id,
                            change_summary=f"+1 day for holiday on off day ({day})",
                            full_diff={"bonus_day": str(day)},
                            conn=conn
                        )

        # === Finalize log ===
        c.execute("UPDATE LeaveUpdateLog SET last_updated = ?, march_processed = ?",
                  (today.isoformat(), 1 if crossed_march else march_transfer_done))
        conn.commit()




# --- Utility Functions ---


def load_holiday_json():
    if not os.path.exists(HOLIDAYS_JSON):
        data = {"2025": {}, "defaults": {}}
        with open(HOLIDAYS_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        return data
    with open(HOLIDAYS_JSON, "r", encoding="utf-8") as f:
        return json.load(f)

def add_holiday(date_str: str, name: str):
    data = load_holiday_json()
    year = date_str.split("-")[0]
    data.setdefault(year, {})[date_str] = name
    with open(HOLIDAYS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    audit("Add Holiday", performed_by="system", change_summary=f"{date_str}->{name}")

def remove_holiday(date_str: str):
    data = load_holiday_json()
    year = date_str.split("-")[0]
    if year in data and date_str in data[year]:
        del data[year][date_str]
        with open(HOLIDAYS_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        audit("Remove Holiday", performed_by="system", change_summary=f"{date_str} removed")

def get_calendar_events(username: str=None, year: int=None):
    events = []
    holidays = load_holiday_json()
    years = [str(year)] if year else list(holidays.keys())
    for y in years:
        for d, name in holidays.get(y, {}).items():
            events.append({"title": name, "start": d, "allDay": True, "className": "holiday"})
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        if username:
            c.execute("SELECT u.username, lr.leave_type, lr.start_date, lr.num_days FROM LeaveRequests lr JOIN Users u ON lr.user_id = u.id WHERE u.username = ? AND lr.status = 'Approved'", (username,))
        else:
            c.execute("SELECT u.username, lr.leave_type, lr.start_date, lr.num_days FROM LeaveRequests lr JOIN Users u ON lr.user_id = u.id WHERE lr.status = 'Approved'")
        rows = c.fetchall()
        for r in rows:
            try:
                start_date = datetime.strptime(r["start_date"], "%Y-%m-%d").date()
            except:
                continue
            for i in range(int(max(1, round(float(r["num_days"] or 1))))):
                d = start_date + timedelta(days=i)
                events.append({"title": f"{r['username']}: {r['leave_type']}", "start": d.isoformat(), "allDay": True, "className": "approved"})
    return events

def get_attachment_path_for_leave(leave_id: int):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT attachment_image FROM LeaveRequests WHERE id = ?", (leave_id,))
        r = c.fetchone()
        if not r: return None
        return r["attachment_image"]

def read_attachment_blob(leave_id: int):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT attachment_image FROM LeaveRequests WHERE id = ?", (leave_id,))
        r = c.fetchone()
        if not r or r[0] is None: return None
        return r[0]
    
    
# leave_api.py (NEW VERSION START HERE)
# Backend logic for the Leave Management System

import sqlite3
import hashlib
import json
import configparser
import os
import sys
import re
import shutil
from pathlib import Path
from typing import Dict, List, Union, Optional, Any
from datetime import date, datetime, timedelta
from calendar import monthrange

# Third-party libraries you will need to install:
# pip install openpyxl
from openpyxl import load_workbook


# ==============================================================================
# 1. CONFIGURATION & CONSTANTS
# ==============================================================================

def create_default_config(path="config.ini"):
    """Creates a default config.ini file if it doesn't exist."""
    if not os.path.exists(path):
        config = configparser.ConfigParser()
        config["database"] = {"path": "leave_data.sqlite"}
        config["files"] = {
            "holidays_json": "holidays.json",
            "template_path": os.path.join("templates", "Leave_Application_Form.xlsx"),
            "output_path": os.path.join("output", "Filled_Form.xlsx")
        }
        # Ensure parent directories exist
        os.makedirs("templates", exist_ok=True)
        os.makedirs("output", exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            config.write(f)
        print(f"Created default {path}")

def create_default_holidays(holidays_path):
    """Creates a default holidays.json file if it doesn't exist."""
    if not os.path.exists(holidays_path):
        os.makedirs(os.path.dirname(holidays_path), exist_ok=True)
        holidays = {
            "2025": {"2025-01-01": "New Year's Day"},
            "defaults": {}
        }
        with open(holidays_path, "w", encoding="utf-8") as f:
            json.dump(holidays, f, indent=4)
        print(f"Created default holidays at {holidays_path}")

# --- Determine Base Directory ---
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(os.path.dirname(sys.executable))
else:
    BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__)))

# --- Read Configuration ---
CONFIG_PATH = BASE_DIR / 'config.ini'
create_default_config(CONFIG_PATH) # Ensure it exists

config = configparser.ConfigParser()
config.read(CONFIG_PATH)

# --- Define Constants from Config ---
DB_PATH = BASE_DIR / config.get('database', 'path', fallback='leave_data.sqlite')
HOLIDAYS_JSON = BASE_DIR / config.get('files', 'holidays_json', fallback='holidays.json')
TEMPLATE_PATH = BASE_DIR / config.get('files', 'template_path', fallback=os.path.join('templates', 'Leave_Application_Form.xlsx'))
OUTPUT_PATH = BASE_DIR / config.get('files', 'output_path', fallback=os.path.join('output', 'Filled_Form.xlsx'))

create_default_holidays(HOLIDAYS_JSON) # Ensure it exists

# Canonical mapping from UI labels to Database values
UI_TO_DB = {
    "annual leave": "Annual Leave",
    "emergency leave": "Emergency Leave",
    "sick leave": "Sick Leave",
    "compassionate leave": "Compassionate Leave",
    "working on off/ph/ot": "Working on Off/PH/OT",
    "hospital leave": "Hospital Leave",
    "maternity leave": "Maternity Leave",
    "cultivation leave": "Cultivation Leave",
}


# ==============================================================================
# 2. DATABASE SETUP
# ==============================================================================

def init_db():
    """
    Initializes all database tables with the full schema.
    This is safe to run multiple times.
    """
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        
        # --- Users Table ---
        c.execute('''
            CREATE TABLE IF NOT EXISTS Users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                login_name TEXT UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT CHECK(role IN ('user','admin', 'superadmin', 'it')) NOT NULL DEFAULT 'user',
                address TEXT,
                phone TEXT,
                rest_days_odd TEXT,
                rest_days_even TEXT,
                TotalLeave REAL DEFAULT 0,
                SickLeave REAL DEFAULT 0,
                cultivationLeave REAL DEFAULT 0,
                hospital_leave REAL DEFAULT 0,
                compassionateLeave REAL DEFAULT 0,
                pregnantLeave REAL DEFAULT 0,
                ReplacementLeave REAL DEFAULT 0,
                years_worked INTEGER DEFAULT 0,
                cf_leave REAL DEFAULT 0,
                expired REAL DEFAULT 0,
                upd TEXT -- Last seen update date
            )
        ''')

        # --- LeaveRequests Table ---
        c.execute('''
            CREATE TABLE IF NOT EXISTS LeaveRequests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                leave_type TEXT NOT NULL,
                start_date TEXT NOT NULL,
                num_days REAL NOT NULL,
                notes TEXT,
                status TEXT DEFAULT 'Pending',
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                replacement_leave TEXT,
                attachment_image BLOB,
                bal REAL,
                leave_source TEXT,
                FOREIGN KEY(user_id) REFERENCES Users(id)
            )
        ''')
        
        # --- Audit Log Table ---
        c.execute('''
            CREATE TABLE IF NOT EXISTS LeaveAudit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT,
                performed_by TEXT,
                target_user TEXT,
                target_request_id INTEGER,
                change_summary TEXT,
                full_diff TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # --- System Update Log ---
        c.execute('''
            CREATE TABLE IF NOT EXISTS LeaveUpdateLog (
                last_updated TEXT,
                march_processed INTEGER DEFAULT 0
            )
        ''')
        
        # --- Yearly Snapshot Table ---
        c.execute('''
            CREATE TABLE IF NOT EXISTS LeaveSnapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                snapshot_date DATE DEFAULT CURRENT_DATE,
                TotalLeave REAL,
                SickLeave REAL,
                cultivationLeave REAL,
                compassionateLeave REAL,
                hospital_leave REAL,
                ReplacementLeave REAL,
                pregnantLeave REAL,
                cf_leave REAL,
                FOREIGN KEY(user_id) REFERENCES Users(id)
            )
        ''')
        
        conn.commit()
        print("Database initialized successfully.")

# ==============================================================================
# 3. CORE UTILITY FUNCTIONS
# ==============================================================================

def hash_password(pw: str) -> str:
    """Hashes a password using SHA256."""
    return hashlib.sha256(pw.encode()).hexdigest()

def get_dict_diff(old: dict, new: dict) -> dict:
    """Compares two dictionaries and returns a summary of differences."""
    diff = {}
    for k in new:
        old_val = old.get(k)
        new_val = new[k]
        if str(old_val) != str(new_val):
            diff[k] = [str(old_val), str(new_val)]
    return diff

def log_action(action: str, performed_by: str, target_user: Optional[str] = None,
               target_request_id: Optional[int] = None, change_summary: str = "",
               full_diff: Optional[Union[dict, str]] = None, conn: Optional[sqlite3.Connection] = None):
    """Logs an action to the LeaveAudit table."""
    if isinstance(full_diff, dict):
        full_diff = json.dumps(full_diff, ensure_ascii=False, indent=2)

    sql = """
        INSERT INTO LeaveAudit (action, performed_by, target_user, target_request_id, change_summary, full_diff)
        VALUES (?, ?, ?, ?, ?, ?)
    """
    params = (action, performed_by, target_user, target_request_id, change_summary, full_diff)

    # Use existing connection or create a new one
    if conn:
        conn.execute(sql, params)
    else:
        with sqlite3.connect(DB_PATH, timeout=10) as new_conn:
            new_conn.execute(sql, params)
            new_conn.commit()
            
def normalize_leave_type(label: str) -> str:
    """
    Normalizes a UI label (e.g., 'Annual Leave 年假') to a lowercase key ('annual leave').
    """
    if not label:
        return ""
    s = label.strip()
    # Find the first Chinese character and slice the string before it
    for i, char in enumerate(s):
        if '\u4e00' <= char <= '\u9fff':
            s = s[:i]
            break
    return s.strip().lower()

def load_holiday_json() -> dict:
    """Loads the holidays.json file."""
    if not os.path.exists(HOLIDAYS_JSON):
        create_default_holidays(HOLIDAYS_JSON)
    
    try:
        with open(HOLIDAYS_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading JSON: {e}")
        return {"2025": {}, "defaults": {}}


# ==============================================================================
# 4. BUSINESS LOGIC & DATA MANIPULATION
# ==============================================================================

def verify_login(login_name: str, password: str) -> Optional[Dict[str, Any]]:
    """
    Authenticates a user and returns their full user data as a dictionary if successful.
    Also creates a default admin if none exists.
    """
    pwd_hash = hash_password(password)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        c.execute("SELECT * FROM Users WHERE login_name = ?", (login_name,))
        user_row = c.fetchone()

        if user_row and pwd_hash == user_row["password_hash"]:
            return dict(user_row)

        # Check for and create default admin if necessary
        c.execute("SELECT 1 FROM Users WHERE role = 'admin'")
        if not c.fetchone():
            default_admin_hash = hash_password('admin')
            c.execute("""
                INSERT INTO Users (username, login_name, password_hash, role)
                VALUES ('admin', 'admin', ?, 'admin')
            """, (default_admin_hash,))
            conn.commit()
            print("Default admin created. Username: 'admin', Password: 'admin'")
            # Try to log in as default admin after creation
            if login_name == 'admin' and password == 'admin':
                 c.execute("SELECT * FROM Users WHERE login_name = 'admin'")
                 return dict(c.fetchone())

    return None

def get_alternating_rest_days(username: str, target_date: date) -> List[str]:
    """
    Returns a list of rest days for a user on a specific date,
    considering odd/even weeks.
    """
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT rest_days_odd, rest_days_even FROM Users WHERE username = ?", (username,))
        row = c.fetchone()

    if not row:
        return []

    rest_days_odd, rest_days_even = row
    week_num = target_date.isocalendar()[1]
    rest_days_str = rest_days_odd if week_num % 2 == 1 else rest_days_even
    return [d.strip().lower() for d in (rest_days_str or "").split(",") if d.strip()]

def calculate_entitlement(leave_type: str, years_worked: int) -> float:
    """Calculates the yearly entitlement for a given leave type."""
    lt = normalize_leave_type(leave_type)
    if lt in ["sick leave", "mc"]:
        return 18.0 if years_worked >= 5 else 12.0
    elif lt in ["hospital leave", "hospital"]:
        return 60.0
    elif lt in ["cultivation leave", "修行假"]:
        return 7.0
    elif lt in ["compassionate leave"]:
        return 14.0
    elif lt in ["annual leave", "emergency leave"]:
        if years_worked > 10:
            return 19.0
        elif years_worked > 5:
            return 18.0
        else:
            return 12.0
    return 0.0

def _deduct_from_column(c: sqlite3.Cursor, user_id: int, column: str, need: float, sources: list, label: str) -> float:
    """Helper function to deduct leave from a specific column."""
    c.execute(f"SELECT {column} FROM Users WHERE id = ?", (user_id,))
    available = float(c.fetchone()[0] or 0.0)
    if available <= 0 or need <= 0:
        return need
    use = min(available, need)
    c.execute(f"UPDATE Users SET {column} = {column} - ? WHERE id = ?", (use, user_id))
    sources.append((label, use))
    return need - use

def deduct_annual_leave(c: sqlite3.Cursor, user_id: int, num_days: float, ref_date: Optional[date] = None) -> str:
    """
    Deducts annual/emergency leave based on seasonal priority.
    - Jan-Mar: cf_leave -> ReplacementLeave -> TotalLeave
    - Apr-Dec: ReplacementLeave -> cf_leave -> TotalLeave
    Returns a string summarizing the deduction sources.
    """
    if ref_date is None:
        ref_date = date.today()
    month = ref_date.month
    use_cf_first = month <= 3

    sources: List[tuple[str, float]] = []
    
    if use_cf_first:
        need = _deduct_from_column(c, user_id, "cf_leave", num_days, sources, "CF")
        need = _deduct_from_column(c, user_id, "ReplacementLeave", need, sources, "Replacement")
        need = _deduct_from_column(c, user_id, "TotalLeave", need, sources, "Annual")
    else:
        need = _deduct_from_column(c, user_id, "ReplacementLeave", num_days, sources, "Replacement")
        need = _deduct_from_column(c, user_id, "cf_leave", need, sources, "CF")
        need = _deduct_from_column(c, user_id, "TotalLeave", need, sources, "Annual")

    if need > 1e-9:
        raise ValueError(f"Insufficient annual/emergency leave balance. Short by {need:.2f} day(s).")

    return ", ".join(f"{src}:{amt:.1f}" for src, amt in sources)

def refund_annual_leave(c: sqlite3.Cursor, user_id: int, leave_source: str):
    """Refunds annual leave based on a source string (e.g., 'CF:1.0, Annual:2.0')."""
    if not leave_source:
        return
    parts = [p.strip() for p in leave_source.split(",") if p.strip()]
    for part in parts:
        if ":" not in part:
            continue
        src, amt_str = part.split(":")
        try:
            amt = float(amt_str)
            col_map = {"CF": "cf_leave", "Replacement": "ReplacementLeave", "Annual": "TotalLeave"}
            column = col_map.get(src)
            if column:
                c.execute(f"UPDATE Users SET {column} = {column} + ? WHERE id = ?", (amt, user_id))
        except (ValueError, KeyError):
            print(f"Warning: Could not parse or refund leave source part: {part}")
            
def apply_annual_leave_updates():
    """
    SYSTEM FUNCTION: To be run periodically (e.g., via a scheduled task or on app startup).
    Handles monthly top-ups, yearly resets, and CF leave expiry.
    """
    today = date.today()

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        conn.row_factory = sqlite3.Row

        # Get system log state
        log_row = c.execute("SELECT last_updated, march_processed FROM LeaveUpdateLog").fetchone()
        if log_row:
            last_updated = date.fromisoformat(log_row['last_updated'])
            march_transfer_done = bool(log_row['march_processed'])
        else:
            last_updated = date(2000, 1, 1)
            march_transfer_done = False
            c.execute("INSERT INTO LeaveUpdateLog (last_updated, march_processed) VALUES (?, 0)", (last_updated.isoformat(),))

        # Skip if already updated this month
        if last_updated.year == today.year and last_updated.month == today.month:
            print("Leave already updated this month.")
            return

        is_new_year = today.year > last_updated.year
        crossed_march = today.month > 3 and not march_transfer_done
        
        # Yearly reset needs to happen in January
        if is_new_year and today.month == 1:
            users_for_reset = c.execute("SELECT id, years_worked, TotalLeave FROM Users").fetchall()
            for user_id, years_worked, current_total in users_for_reset:
                # Logic for yearly reset
                new_sick_leave = 18 if years_worked + 1 > 5 else 12
                new_cf = min(current_total, 5)
                c.execute("""
                    UPDATE Users SET
                        cf_leave = ?, TotalLeave = 0, SickLeave = ?,
                        hospital_leave = 60, cultivationLeave = 7, compassionateLeave = 14,
                        ReplacementLeave = 0, pregnantLeave = 98,
                        years_worked = years_worked + 1
                    WHERE id = ?
                """, (new_cf, new_sick_leave, user_id))
                log_action("Yearly Reset", "System", target_user=str(user_id),
                           change_summary=f"CF set to {new_cf}, SickLeave to {new_sick_leave}", conn=conn)

        # March CF expiry
        if crossed_march:
            users_for_expiry = c.execute("SELECT id, cf_leave FROM Users WHERE cf_leave > 0").fetchall()
            for user_id, cf_old in users_for_expiry:
                c.execute("UPDATE Users SET cf_leave = 0 WHERE id = ?", (user_id,))
                log_action("March Expiry", "System", target_user=str(user_id),
                           change_summary=f"CF leave of {cf_old} expired.", conn=conn)
        
        # Monthly Top-up
        users_for_topup = c.execute("SELECT id, years_worked, TotalLeave FROM Users").fetchall()
        for user_id, years_worked, current_total in users_for_topup:
            if years_worked > 10:
                add_total = 1.58334
            elif years_worked > 5:
                add_total = 1.5
            else:
                add_total = 1.0
            
            new_total = round(current_total + add_total, 2)
            c.execute("UPDATE Users SET TotalLeave = ? WHERE id = ?", (new_total, user_id))
            log_action("Monthly Top-up", "System", target_user=str(user_id),
                       change_summary=f"Added {add_total}. New TotalLeave: {new_total}", conn=conn)
        
        # Update log
        c.execute("UPDATE LeaveUpdateLog SET last_updated = ?, march_processed = ?",
                  (today.isoformat(), 1 if crossed_march else march_transfer_done))
        
        conn.commit()
        print("Monthly leave updates applied.")

# ==============================================================================
# 5. "API" FUNCTIONS FOR FLASK
# These functions will be called by your Flask routes.
# ==============================================================================

# --- GET (Read) Operations ---

def get_user_profile(user_id: int) -> Optional[Dict[str, Any]]:
    """Fetches a single user's complete profile and leave balances."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        user = conn.execute("SELECT * FROM Users WHERE id = ?", (user_id,)).fetchone()
        return dict(user) if user else None

def get_all_users() -> List[Dict[str, Any]]:
    """Fetches a list of all users with basic info."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        users = conn.execute("SELECT id, username, login_name, role FROM Users ORDER BY username").fetchall()
        return [dict(u) for u in users]

def get_user_leave_requests(user_id: int) -> List[Dict[str, Any]]:
    """Fetches all leave requests for a specific user."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        requests = conn.execute(
            "SELECT * FROM LeaveRequests WHERE user_id = ? ORDER BY start_date DESC",
            (user_id,)
        ).fetchall()
        return [dict(r) for r in requests]

def get_leave_request_by_id(request_id: int) -> Optional[Dict[str, Any]]:
    """Fetches a single leave request by its ID."""
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        request = conn.execute("SELECT * FROM LeaveRequests WHERE id = ?", (request_id,)).fetchone()
        return dict(request) if request else None
        
def get_holidays_for_year(year: int) -> Dict[str, str]:
    """Returns a flat dictionary of holidays for a given year."""
    all_holidays = {}
    holidays_data = load_holiday_json()
    year_str = str(year)

    # Add default holidays for the given year
    for mm_dd, name in holidays_data.get("defaults", {}).items():
        all_holidays[f"{year_str}-{mm_dd}"] = name

    # Add/overwrite with year-specific holidays
    for yyyy_mm_dd, name in holidays_data.get(year_str, {}).items():
        all_holidays[yyyy_mm_dd] = name
        
    return all_holidays

def get_pending_requests(approver_id: int) -> List[Dict[str, Any]]:
    """
    Fetches pending leave requests based on the approver's role.
    This encapsulates the complex logic from the AdminApprovalTab.
    """
    approver = get_user_profile(approver_id)
    if not approver:
        return []

    role = approver['role']
    query = """
        SELECT lr.*, u.username, u.role as user_role
        FROM LeaveRequests lr JOIN Users u ON lr.user_id = u.id
        WHERE 1=1
    """
    params = []

    if role.startswith("admin") and len(role) == 6:  # e.g., adminA
        role_prefix = role[-1].upper()
        query += " AND u.role = ? AND lr.status = 'Pending' AND u.role NOT LIKE 'admin%' AND u.role != 'superadmin'"
        params.append(f"user{role_prefix}")
    elif role == "admin":
        query += " AND u.role != 'admin' AND lr.status IN ('Pending', 'Under Approval')"
    elif role == "superadmin":
        query += " AND u.role = 'admin' AND lr.status IN ('Pending', 'Under Approval')"
    elif role == "it":
        query += " AND lr.status IN ('Pending', 'Under Approval')"
    else:
        return [] # Regular users can't see this

    query += " ORDER BY lr.timestamp DESC"

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        requests = conn.execute(query, params).fetchall()
        return [dict(r) for r in requests]

# --- POST/PUT (Write) Operations ---

def create_leave_request(user_id: int, leave_type: str, selected_dates: Dict[date, float],
                         notes: str, attachment_blob: Optional[bytes] = None,
                         replacement_info: Optional[str] = None) -> None:
    """
    Creates one or more leave requests based on user input.
    Handles date expansion, OFF/PH filtering, and batching via a random code.
    """
    user = get_user_profile(user_id)
    if not user:
        raise ValueError(f"User with ID {user_id} not found.")

    username = user['username']
    leave_type_db = UI_TO_DB.get(normalize_leave_type(leave_type), leave_type)
    
    # Generate a batch code for this submission
    batch_code = f"[{datetime.now().strftime('%f')}{user_id}]"
    notes_with_code = f"{batch_code} {notes.strip()}"

    # Expand multi-day requests into individual entries
    expanded_dates = []
    for d, duration in selected_dates.items():
        days_to_add = int(duration)
        for i in range(days_to_add):
            expanded_dates.append(d + timedelta(days=i))
        if duration % 1 != 0: # Handle half days
            expanded_dates.append(d + timedelta(days=days_to_add))
    
    # Filter out OFF days and Public Holidays unless it's a "Working on Off Day" request
    final_dates = []
    if normalize_leave_type(leave_type) != "working on off/ph/ot":
        holidays = get_holidays_for_year(date.today().year) # Simple: just checks current year
        for d in expanded_dates:
            rest_days = get_alternating_rest_days(username, d)
            is_rest_day = d.strftime('%A').lower() in rest_days
            is_holiday = d.isoformat() in holidays
            if not is_rest_day and not is_holiday:
                final_dates.append(d)
    else:
        final_dates = expanded_dates

    if not final_dates:
        raise ValueError("No valid working days selected for leave.")

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        for d in final_dates:
            num_days_for_date = 1.0 # Simplified for API
            c.execute("""
                INSERT INTO LeaveRequests
                    (user_id, leave_type, start_date, num_days, status, notes, replacement_leave, attachment_image)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                user_id, leave_type_db, d.isoformat(), num_days_for_date, "Pending",
                notes_with_code, replacement_info, attachment_blob
            ))
        conn.commit()

def approve_request(request_id: int, approver_id: int):
    """
    Approves a leave request, deducts balance, and logs the action.
    This function also handles batch approval based on the note's batch code.
    """
    approver = get_user_profile(approver_id)
    if not approver:
        raise PermissionError("Approver not found.")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # Get the primary request to find the batch code
        main_req = c.execute("SELECT notes FROM LeaveRequests WHERE id = ?", (request_id,)).fetchone()
        if not main_req or not main_req['notes']:
            raise ValueError("Request or batch code not found.")
        
        match = re.search(r"(\[\d+\])", main_req['notes'])
        if not match:
            raise ValueError("Invalid batch code format in notes.")
        batch_code = match.group(1)

        # Get all requests in the batch
        requests_to_process = c.execute(
            "SELECT * FROM LeaveRequests WHERE notes LIKE ?",
            (f"%{batch_code}%",)
        ).fetchall()

        try:
            for req in requests_to_process:
                # This is a simplified approval logic. The original was very complex.
                # A real-world scenario would replicate the full logic from _approve_request_by_id here.
                # For now, we'll do a basic approval and deduction.
                
                leave_type = normalize_leave_type(req['leave_type'])
                num_days = req['num_days']
                user_id = req['user_id']
                
                if leave_type == "working on off/ph/ot":
                    c.execute("UPDATE Users SET ReplacementLeave = ReplacementLeave + ? WHERE id = ?", (num_days, user_id))
                elif leave_type in ["annual leave", "emergency leave"]:
                    deduct_annual_leave(c, user_id, num_days, date.fromisoformat(req['start_date']))
                else:
                    # Simple deduction for other types
                    col_map = {
                        "sick leave": "SickLeave", "compassionate leave": "compassionateLeave",
                        "hospital leave": "hospital_leave", "maternity leave": "pregnantLeave",
                        "cultivation leave": "cultivationLeave"
                    }
                    column = col_map.get(leave_type)
                    if column:
                        c.execute(f"UPDATE Users SET {column} = {column} - ? WHERE id = ?", (num_days, user_id))
                
                # Update status
                c.execute("UPDATE LeaveRequests SET status = 'Approved' WHERE id = ?", (req['id'],))
                log_action("Approve", approver['username'], target_request_id=req['id'], conn=conn)

            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e # Re-raise the exception for Flask to handle

def reject_request(request_id: int, approver_id: int):
    """Rejects a leave request and logs the action, handling batches."""
    approver = get_user_profile(approver_id)
    if not approver:
        raise PermissionError("Approver not found.")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        main_req = c.execute("SELECT notes FROM LeaveRequests WHERE id = ?", (request_id,)).fetchone()
        if not main_req or not main_req['notes']:
            raise ValueError("Request or batch code not found.")
        
        match = re.search(r"(\[\d+\])", main_req['notes'])
        if not match:
            raise ValueError("Invalid batch code format in notes.")
        batch_code = match.group(1)

        requests_to_process = c.execute(
            "SELECT * FROM LeaveRequests WHERE notes LIKE ?",
            (f"%{batch_code}%",)
        ).fetchall()
        
        try:
            for req in requests_to_process:
                # If it was previously approved, refund the leave
                if req['status'] == 'Approved' and req['leave_source']:
                    if ":" in req['leave_source']:
                        refund_annual_leave(c, req['user_id'], req['leave_source'])
                    else: # Simple column refund
                        c.execute(f"UPDATE Users SET {req['leave_source']} = {req['leave_source']} + ? WHERE id = ?",
                                  (req['num_days'], req['user_id']))

                c.execute("UPDATE LeaveRequests SET status = 'Rejected' WHERE id = ?", (req['id'],))
                log_action("Reject", approver['username'], target_request_id=req['id'], conn=conn)

            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

def create_or_update_user(user_data: Dict[str, Any], performed_by_id: int) -> int:
    """Creates a new user or updates an existing one."""
    admin = get_user_profile(performed_by_id)
    if not admin or admin['role'] not in ['admin', 'it', 'superadmin']:
        raise PermissionError("You do not have permission to manage users.")

    user_id = user_data.get('id')
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        if user_id: # Update
            # In a real app, you'd fetch the old data for logging the diff
            sql = """UPDATE Users SET username=?, login_name=?, role=?, address=?, phone=? WHERE id=?"""
            params = (user_data['username'], user_data['login_name'], user_data['role'], 
                      user_data['address'], user_data['phone'], user_id)
            c.execute(sql, params)
        else: # Create
            password_hash = hash_password(user_data['password'])
            sql = """INSERT INTO Users (username, login_name, password_hash, role) VALUES (?, ?, ?, ?)"""
            params = (user_data['username'], user_data['login_name'], password_hash, user_data['role'])
            c.execute(sql, params)
            user_id = c.lastrowid
        conn.commit()
    return user_id

# ==============================================================================
# 6. SERVICE FUNCTIONS (e.g., Excel Export)
# ==============================================================================

def export_leave_to_excel(leave_id: int) -> str:
    """
    Fills an Excel template with leave data and returns the path to the output file.
    """
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        leave = c.execute("SELECT * FROM LeaveRequests WHERE id = ?", (leave_id,)).fetchone()
        if not leave:
            raise ValueError("Leave request not found")

        user = c.execute("SELECT * FROM Users WHERE id = ?", (leave["user_id"],)).fetchone()
        if not user:
            raise ValueError("User not found")

    # --- Calculations & Data Preparation ---
    clean_notes = re.sub(r"\[\d+\]\s*", "", leave["notes"] or "")
    num_days = leave["num_days"]
    leave_type = leave["leave_type"]
    is_basic = normalize_leave_type(leave_type) in ["annual leave", "emergency leave", "working on off/ph/ot"]

    cf = user["cf_leave"] if is_basic else 0
    rep = user["ReplacementLeave"] if is_basic else 0
    ent = user["TotalLeave"] if is_basic else 0 # Simplified: should be yearly entitlement
    
    # --- Start Filling Excel ---
    output_file_path = OUTPUT_PATH.parent / f"Leave_Form_{leave_id}_{user['username']}.xlsx"
    shutil.copyfile(TEMPLATE_PATH, output_file_path)
    wb = load_workbook(output_file_path)
    ws = wb.active

    ws["D13"] = user["username"]
    ws["L13"] = f"Tel: {user['phone'] or '-'}"
    ws["D15"] = user["address"] or "-"
    ws["E31"] = clean_notes
    ws["E33"] = leave["start_date"]
    ws["M31"] = str(num_days)
    ws["E35"] = str(num_days)
    ws["B40"] = datetime.strptime(leave["timestamp"], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
    ws["M19"] = str(ent)
    ws["M21"] = str(cf)
    
    wb.save(output_file_path)
    return str(output_file_path)

# ==============================================================================
# 7. MAIN EXECUTION BLOCK (for setup)
# ==============================================================================

if __name__ == '__main__':
    print("Running initial setup for leave_api.py...")
    
    # 1. Ensure config and holiday files exist
    create_default_config(CONFIG_PATH)
    create_default_holidays(HOLIDAYS_JSON)
    
    # 2. Initialize the database schema
    init_db()
    
    # 3. Create a default admin user if one doesn't exist
    verify_login('nonexistentuser', 'fakepass') # This triggers the admin check
    
    print("\nSetup complete. You can now import and use these functions in your Flask app.")
    
    # Example Usage (demonstration)
    print("\n--- Example Usage ---")
    try:
        admin = verify_login('admin', 'admin')
        if admin:
            print("Successfully logged in as admin.")
            pending = get_pending_requests(admin['id'])
            print(f"Found {len(pending)} pending requests for admin.")
        else:
            print("Could not log in as admin. Check password.")
            
        users = get_all_users()
        print(f"Total users in system: {len(users)}")
        if users:
            print(f"First user: {users[0]['username']}")
            
    except Exception as e:
        print(f"An error occurred during example usage: {e}")

